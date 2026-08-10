"""Build Overview KPI Excel calculator template (Excel 365)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "templates" / "Overview_KPI_Calculator_Template.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1B7A4E")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
TITLE_FONT = Font(bold=True, name="Calibri", size=16, color="1B7A4E")
SECTION_FONT = Font(bold=True, name="Calibri", size=12, color="1B7A4E")
LABEL_FONT = Font(bold=True, name="Calibri", size=11)
HINT_FONT = Font(italic=True, name="Calibri", size=10, color="666666")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
INPUT_FILL = PatternFill("solid", fgColor="FFF8E7")
KPI_FILL = PatternFill("solid", fgColor="E8F5EE")
WARN_FILL = PatternFill("solid", fgColor="FFF3CD")

# Pre-fill formula rows so UNIQUE spills have adjacent measures ready
TRIP_ROWS = 2500
DAILY_ROWS = 100
ROUTE_ROWS = 150


def style_header_row(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN


def autosize(ws, widths: dict[int, float]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True
    )
    ws.add_table(table)


def build_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "E-TRAM Overview KPI Calculator — Excel template"
    ws["A1"].font = TITLE_FONT
    lines = [
        "",
        "Requires Microsoft Excel 365 (UNIQUE, FILTER, XLOOKUP, TEXTAFTER).",
        "",
        "How to use",
        "1. Open KPI_Controls — set StartDate, EndDate, and optional RouteFilter.",
        "2. Paste Veh_Type into Veh (keep headers). VehClean / formulas auto-fill in the table.",
        "3. Paste Route_Description into Routes. RDK formula auto-fills.",
        "4. Paste Bus Ticket Data into Tickets starting at row 2 under the existing headers.",
        "   Helper columns TP…RouteLen are already formula columns — do not overwrite them.",
        "5. TripSummary!A2 spills unique trip keys for the selected period.",
        "   Columns B–L are pre-filled for up to {trip} trips.",
        "6. Daily and RouteDay spill from TripSummary; KPI sheet shows Overview totals.",
        "",
        "Sheet map",
        "KPI_Controls — date / route filter (named ranges StartDate, EndDate, RouteFilter)",
        "Tickets — ETM ticket rows + helper columns",
        "Veh — vehicle capacity lookup",
        "Routes — route length lookup",
        "TripSummary — one row per bus trip in period",
        "Daily — agency-day aggregates",
        "RouteDay — route contribution / shares",
        "KPI — Overview primary cards + operating snapshot",
        "",
        "Notes",
        "• Load factor = SUM(PaxKm) / SUM(CapacityKm), never an average of daily ratios.",
        "• Trips per bus = SUM(trips) / SUM(bus-days), not trips ÷ average buses.",
        "• Capacity and route length must resolve; otherwise LF / EPKM stay blank or zero.",
        "• Strip spaces from vehicle IDs (done via VehClean) to match Supporting data.",
        "• HeadwayMins ≈ (max trip start − min trip start) × 1440 / trips — dashboard approximation.",
        "",
        "Source headers (Bus Ticket Data)",
        "Date | Ticket No. | Route No. | Route description | Depot | Vehicle/ Schedule no. |",
        "Driver ID | Conductor ID | Trip No. | Trip Start Time | Trip End Time | Ticket Issue Time |",
        "No. of pass. | No. of child Pass. | Pass. Category | Origin Stop No. | Destination Stop No. |",
        "Pass. Origin | Pass. Destination | Stage Km | Revenue | Gender",
    ]
    for i, line in enumerate(lines, start=2):
        ws.cell(i, 1, line.format(trip=TRIP_ROWS))
        if line in {
            "How to use",
            "Sheet map",
            "Notes",
            "Source headers (Bus Ticket Data)",
        }:
            ws.cell(i, 1).font = SECTION_FONT
    autosize(ws, {1: 110})


def build_controls(wb: Workbook) -> None:
    ws = wb.create_sheet("KPI_Controls")
    ws["A1"] = "Overview period filters"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = "StartDate"
    ws["A4"] = "EndDate"
    ws["A5"] = "RouteFilter"
    ws["B3"] = "2025-04-01"
    ws["B4"] = "2025-04-30"
    ws["B5"] = ""
    for r in (3, 4, 5):
        ws.cell(r, 1).font = LABEL_FONT
        ws.cell(r, 2).fill = INPUT_FILL
        ws.cell(r, 2).border = THIN
    ws["C3"] = "Inclusive start (date)"
    ws["C4"] = "Inclusive end (date)"
    ws["C5"] = "Leave blank for all routes, or enter one Route No. (e.g. 1)"
    for r in (3, 4, 5):
        ws.cell(r, 3).font = HINT_FONT
    ws["A7"] = "Named ranges StartDate / EndDate / RouteFilter point at B3:B5."
    ws["A7"].font = HINT_FONT
    ws["A9"] = "After pasting tickets, confirm KPI!B5 (ridership) updates."
    ws["A9"].font = HINT_FONT
    autosize(ws, {1: 16, 2: 16, 3: 70})

    wb.defined_names.add(DefinedName(name="StartDate", attr_text="KPI_Controls!$B$3"))
    wb.defined_names.add(DefinedName(name="EndDate", attr_text="KPI_Controls!$B$4"))
    wb.defined_names.add(DefinedName(name="RouteFilter", attr_text="KPI_Controls!$B$5"))


def build_veh(wb: Workbook) -> None:
    ws = wb.create_sheet("Veh")
    headers = ["Vehicle no./ Schedule No.", "Veh. Type", "Veh. Capacity", "VehClean", "Capacity"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header_row(ws, 1, len(headers))
    # One example + blank template rows inside table
    ws["A2"] = "GJ04XX0001"
    ws["B2"] = "Standard"
    ws["C2"] = 40
    ws["D2"] = '=SUBSTITUTE([@[Vehicle no./ Schedule No.]]," ","")'
    ws["E2"] = "=[@[Veh. Capacity]]"
    for r in range(3, 6):
        ws.cell(r, 4, '=SUBSTITUTE([@[Vehicle no./ Schedule No.]]," ","")')
        ws.cell(r, 5, "=[@[Veh. Capacity]]")
    add_table(ws, "Veh", "A1:E5")
    autosize(ws, {1: 28, 2: 14, 3: 14, 4: 18, 5: 12})
    ws["A7"] = "Replace sample row; paste Veh_Type under headers. Keep VehClean and Capacity formula columns."
    ws["A7"].font = HINT_FONT


def build_routes(wb: Workbook) -> None:
    ws = wb.create_sheet("Routes")
    headers = [
        "Route_Code",
        "Route",
        "Route_description",
        "Route Length",
        "Route Category",
        "RDK",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header_row(ws, 1, len(headers))
    ws["A2"] = "1"
    ws["B2"] = "Sample route"
    ws["C2"] = "A to B"
    ws["D2"] = 12.5
    ws["E2"] = "City"
    ws["F2"] = '=[@Route_Code]&"-"&[@Route_description]'
    for r in range(3, 6):
        ws.cell(r, 6, '=[@Route_Code]&"-"&[@Route_description]')
    add_table(ws, "Routes", "A1:F5")
    autosize(ws, {1: 12, 2: 18, 3: 28, 4: 14, 5: 14, 6: 28})
    ws["A7"] = "Paste Route_Description. RDK must match Tickets RDK exactly (code + description)."
    ws["A7"].font = HINT_FONT


def build_tickets(wb: Workbook) -> None:
    ws = wb.create_sheet("Tickets")
    src = [
        "Date",
        "Ticket No.",
        "Route No.",
        "Route description",
        "Depot",
        "Vehicle/ Schedule no.",
        "Driver ID",
        "Conductor ID",
        "Trip No.",
        "Trip Start Time",
        "Trip End Time",
        "Ticket Issue Time",
        "No. of pass.",
        "No. of child Pass.",
        "Pass. Category",
        "Origin Stop No.",
        "Destination Stop No.",
        "Pass. Origin",
        "Pass. Destination",
        "Stage Km",
        "Revenue",
        "Gender",
    ]
    helpers = [
        "TP",
        "PK",
        "VehClean",
        "BTK",
        "RDK",
        "TripKey",
        "InPeriod",
        "Capacity",
        "RouteLen",
    ]
    headers = src + helpers
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header_row(ws, 1, len(headers))

    # Sample ticket so formulas resolve before real paste
    sample = {
        1: "2025-04-01",
        2: "T001",
        3: "1",
        4: "A to B",
        5: "Depot1",
        6: "GJ04XX0001",
        7: "D1",
        8: "C1",
        9: 1,
        10: "08:00",
        11: "09:00",
        12: "08:15",
        13: 1,
        14: 0,
        15: "Cash",
        16: 1,
        17: 5,
        18: "STOPA",
        19: "STOPB",
        20: 5.0,
        21: 10.0,
        22: "M",
    }
    for c, v in sample.items():
        ws.cell(2, c, v)

    # Helper formulas use structured refs — need table first; openpyxl writes formulas with [@]
    helper_formulas = {
        23: "=[@[No. of pass.]]+[@[No. of child Pass.]]",
        24: "=[@TP]*[@[Stage Km]]",
        25: '=SUBSTITUTE([@[Vehicle/ Schedule no.]]," ","")',
        26: '=[@VehClean]&"-"&[@[Trip No.]]',
        27: '=[@[Route No.]]&"-"&[@[Route description]]',
        28: '=TEXT([@Date],"yyyy-mm-dd")&"|"&[@[Route No.]]&"|"&[@BTK]',
        29: "=AND([@Date]>=StartDate,[@Date]<=EndDate)",
        30: '=IFERROR(XLOOKUP([@VehClean],Veh[VehClean],Veh[Capacity]),0)',
        31: '=IFERROR(XLOOKUP([@RDK],Routes[RDK],Routes[Route Length]),0)',
    }
    # Extra blank rows in table for paste expansion cue
    for r in range(2, 6):
        for c, f in helper_formulas.items():
            ws.cell(r, c, f)

    add_table(ws, "Tickets", f"A1:{get_column_letter(len(headers))}5")
    autosize(ws, {i: 14 for i in range(1, len(headers) + 1)})
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["AB"].width = 28  # TripKey area approx

    ws.cell(7, 1, "Paste real ETM rows under the headers. Delete the sample row after paste.")
    ws.cell(7, 1).font = HINT_FONT
    ws.cell(8, 1, "Do not paste over columns TP onward — those are calculated.")
    ws.cell(8, 1).fill = WARN_FILL


def build_trip_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("TripSummary")
    headers = [
        "TripKey",
        "Date",
        "Route",
        "BTK",
        "Ridership",
        "Revenue",
        "PaxKm",
        "Vehicle",
        "RouteLen",
        "Capacity",
        "CapacityKm",
        "TripStart",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header_row(ws, 1, len(headers))

    # Spill unique keys
    ws["A2"] = (
        '=IFERROR(UNIQUE(FILTER(Tickets[TripKey],'
        "(Tickets[Date]>=StartDate)*(Tickets[Date]<=EndDate)*"
        '(IF(RouteFilter="",TRUE,Tickets[Route No.]=RouteFilter))'
        ')),"")'
    )

    # Pre-fill measure formulas for many rows (not a Table — spill + table conflict)
    for r in range(2, TRIP_ROWS + 2):
        ws.cell(r, 2, f'=IF(A{r}="","",IFERROR(DATEVALUE(TEXTBEFORE(A{r},"|")),--LEFT(A{r},10)))')
        ws.cell(r, 2).number_format = "yyyy-mm-dd"
        ws.cell(r, 3, f'=IF(A{r}="","",TEXTBEFORE(TEXTAFTER(A{r},"|"),"|"))')
        ws.cell(r, 4, f'=IF(A{r}="","",TEXTAFTER(A{r},"|",2))')
        ws.cell(r, 5, f'=IF(A{r}="","",SUMIF(Tickets[TripKey],A{r},Tickets[TP]))')
        ws.cell(r, 6, f'=IF(A{r}="","",SUMIF(Tickets[TripKey],A{r},Tickets[Revenue]))')
        ws.cell(r, 7, f'=IF(A{r}="","",SUMIF(Tickets[TripKey],A{r},Tickets[PK]))')
        ws.cell(r, 8, f'=IF(A{r}="","",IFERROR(XLOOKUP(A{r},Tickets[TripKey],Tickets[VehClean]),""))')
        ws.cell(r, 9, f'=IF(A{r}="","",IFERROR(XLOOKUP(A{r},Tickets[TripKey],Tickets[RouteLen]),0))')
        ws.cell(r, 10, f'=IF(A{r}="","",IFERROR(XLOOKUP(A{r},Tickets[TripKey],Tickets[Capacity]),0))')
        ws.cell(r, 11, f'=IF(A{r}="","",I{r}*J{r})')
        ws.cell(r, 12, f'=IF(A{r}="","",IFERROR(MINIFS(Tickets[Trip Start Time],Tickets[TripKey],A{r}),""))')

    autosize(ws, {1: 32, 2: 12, 3: 10, 4: 16, 5: 11, 6: 11, 7: 11, 8: 14, 9: 10, 10: 10, 11: 12, 12: 12})
    ws.cell(TRIP_ROWS + 3, 1, f"Pre-filled measure formulas through row {TRIP_ROWS + 1}. Extend if you have more trips.")
    ws.cell(TRIP_ROWS + 3, 1).font = HINT_FONT


def build_daily(wb: Workbook) -> None:
    ws = wb.create_sheet("Daily")
    headers = [
        "Date",
        "Ridership",
        "Revenue",
        "PaxKm",
        "CapacityKm",
        "Trips",
        "Buses",
        "LF",
        "ATL",
        "FareYield",
        "TripsPerBus",
        "EPKM",
        "EPB",
        "HeadwayMins",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header_row(ws, 1, len(headers))

    ws["A2"] = '=IFERROR(UNIQUE(FILTER(TripSummary[Date],TripSummary[Date]<>"")),"")'
    # TripSummary is not a table — use column ranges instead for FILTER
    # Fix: use A:A style ranges on TripSummary sheet
    ws["A2"] = (
        '=IFERROR(UNIQUE(FILTER(TripSummary!$B$2:$B$'
        f"{TRIP_ROWS + 1},TripSummary!$B$2:$B${TRIP_ROWS + 1}<>\"\")),\"\")"
    )

    last = TRIP_ROWS + 1
    for r in range(2, DAILY_ROWS + 2):
        ws.cell(r, 2, f'=IF(A{r}="","",SUMIF(TripSummary!$B$2:$B${last},A{r},TripSummary!$E$2:$E${last}))')
        ws.cell(r, 3, f'=IF(A{r}="","",SUMIF(TripSummary!$B$2:$B${last},A{r},TripSummary!$F$2:$F${last}))')
        ws.cell(r, 4, f'=IF(A{r}="","",SUMIF(TripSummary!$B$2:$B${last},A{r},TripSummary!$G$2:$G${last}))')
        ws.cell(r, 5, f'=IF(A{r}="","",SUMIF(TripSummary!$B$2:$B${last},A{r},TripSummary!$K$2:$K${last}))')
        ws.cell(r, 6, f'=IF(A{r}="","",COUNTIF(TripSummary!$B$2:$B${last},A{r}))')
        ws.cell(
            r,
            7,
            f'=IF(A{r}="","",IFERROR(ROWS(UNIQUE(FILTER(TripSummary!$H$2:$H${last},TripSummary!$B$2:$B${last}=A{r}))),0))',
        )
        ws.cell(r, 8, f'=IF(OR(A{r}="",E{r}=0),"" ,D{r}/E{r})')
        ws.cell(r, 9, f'=IF(OR(A{r}="",B{r}=0),"",D{r}/B{r})')
        ws.cell(r, 10, f'=IF(OR(A{r}="",B{r}=0),"",C{r}/B{r})')
        ws.cell(r, 11, f'=IF(OR(A{r}="",G{r}=0),"",F{r}/G{r})')
        ws.cell(
            r,
            12,
            f'=IF(A{r}="","",IFERROR(C{r}/SUMIF(TripSummary!$B$2:$B${last},A{r},TripSummary!$I$2:$I${last}),0))',
        )
        ws.cell(r, 13, f'=IF(OR(A{r}="",G{r}=0),"",C{r}/G{r})')
        ws.cell(
            r,
            14,
            f'=IF(OR(A{r}="",F{r}=0),"",'
            f"(MAXIFS(TripSummary!$L$2:$L${last},TripSummary!$B$2:$B${last},A{r})"
            f"-MINIFS(TripSummary!$L$2:$L${last},TripSummary!$B$2:$B${last},A{r}))*1440/F{r})",
        )

    autosize(ws, {i: 12 for i in range(1, 15)})
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["N"].width = 12


def build_route_day(wb: Workbook) -> None:
    ws = wb.create_sheet("RouteDay")
    headers = [
        "Route",
        "Ridership",
        "Revenue",
        "Trips",
        "PaxKm",
        "CapacityKm",
        "LF",
        "SharePax",
        "ShareRev",
        "FareYield",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
    style_header_row(ws, 1, len(headers))

    last = TRIP_ROWS + 1
    ws["A2"] = (
        f'=IFERROR(UNIQUE(FILTER(TripSummary!$C$2:$C${last},TripSummary!$C$2:$C${last}<>"")),"")'
    )
    for r in range(2, ROUTE_ROWS + 2):
        ws.cell(r, 2, f'=IF(A{r}="","",SUMIF(TripSummary!$C$2:$C${last},A{r},TripSummary!$E$2:$E${last}))')
        ws.cell(r, 3, f'=IF(A{r}="","",SUMIF(TripSummary!$C$2:$C${last},A{r},TripSummary!$F$2:$F${last}))')
        ws.cell(r, 4, f'=IF(A{r}="","",COUNTIF(TripSummary!$C$2:$C${last},A{r}))')
        ws.cell(r, 5, f'=IF(A{r}="","",SUMIF(TripSummary!$C$2:$C${last},A{r},TripSummary!$G$2:$G${last}))')
        ws.cell(r, 6, f'=IF(A{r}="","",SUMIF(TripSummary!$C$2:$C${last},A{r},TripSummary!$K$2:$K${last}))')
        ws.cell(r, 7, f'=IF(OR(A{r}="",F{r}=0),"",E{r}/F{r})')
        ws.cell(r, 8, f'=IF(OR(A{r}="",KPI!$B$5=0),"",B{r}/KPI!$B$5)')
        ws.cell(r, 9, f'=IF(OR(A{r}="",KPI!$B$8=0),"",C{r}/KPI!$B$8)')
        ws.cell(r, 10, f'=IF(OR(A{r}="",B{r}=0),"",C{r}/B{r})')

    for col in ("G", "H", "I"):
        for r in range(2, ROUTE_ROWS + 2):
            ws.cell(r, ord(col) - 64).number_format = "0.0%"
    for r in range(2, ROUTE_ROWS + 2):
        ws.cell(r, 10).number_format = "₹#,##0.00"

    autosize(ws, {1: 10, 2: 12, 3: 12, 4: 10, 5: 12, 6: 12, 7: 10, 8: 10, 9: 10, 10: 12})


def build_kpi(wb: Workbook) -> None:
    ws = wb.create_sheet("KPI")
    ws["A1"] = "Overview KPIs (selected period)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = '=CONCAT("Period: ",TEXT(StartDate,"yyyy-mm-dd")," → ",TEXT(EndDate,"yyyy-mm-dd"),IF(RouteFilter="",""," | Route "&RouteFilter))'
    ws["A2"].font = HINT_FONT

    daily_last = DAILY_ROWS + 1

    def kpi_row(r: int, label: str, formula: str, fmt: str | None = None) -> None:
        ws.cell(r, 1, label).font = LABEL_FONT
        cell = ws.cell(r, 2, formula)
        cell.fill = KPI_FILL
        cell.border = THIN
        if fmt:
            cell.number_format = fmt

    ws["A4"] = "Primary cards"
    ws["A4"].font = SECTION_FONT
    kpi_row(5, "Ridership", f'=SUMIF(Daily!$A$2:$A${daily_last},\"<>\",Daily!$B$2:$B${daily_last})', "#,##0")
    kpi_row(6, "Service days", f'=COUNTIF(Daily!$A$2:$A${daily_last},\"<>\")', "0")
    kpi_row(7, "Passengers / day", "=IF(B6=0,0,B5/B6)", "#,##0")
    kpi_row(8, "Revenue", f'=SUMIF(Daily!$A$2:$A${daily_last},\"<>\",Daily!$C$2:$C${daily_last})', "₹#,##0")
    kpi_row(9, "Revenue / day", "=IF(B6=0,0,B8/B6)", "₹#,##0")
    kpi_row(10, "Fare yield (₹/pax)", "=IF(B5=0,0,B8/B5)", "₹#,##0.00")
    kpi_row(
        11,
        "Load factor (LF)",
        f'=IFERROR(SUMIF(Daily!$A$2:$A${daily_last},"<>",Daily!$D$2:$D${daily_last})/'
        f'SUMIF(Daily!$A$2:$A${daily_last},"<>",Daily!$E$2:$E${daily_last}),0)',
        "0.0%",
    )
    kpi_row(12, "Service trips", f'=SUMIF(Daily!$A$2:$A${daily_last},"<>",Daily!$F$2:$F${daily_last})', "#,##0")
    kpi_row(13, "Buses / day", f'=IFERROR(AVERAGEIF(Daily!$A$2:$A${daily_last},"<>",Daily!$G$2:$G${daily_last}),0)', "0.0")
    kpi_row(
        14,
        "Trips per bus",
        f'=IFERROR(B12/SUMIF(Daily!$A$2:$A${daily_last},"<>",Daily!$G$2:$G${daily_last}),0)',
        "0.0",
    )

    ws["A16"] = "Operating snapshot"
    ws["A16"].font = SECTION_FONT
    kpi_row(17, "ATL (km)", f'=IFERROR(SUMIF(Daily!$A$2:$A${daily_last},"<>",Daily!$D$2:$D${daily_last})/B5,0)', "0.00")
    kpi_row(18, "EPKM (avg of daily)", f'=IFERROR(AVERAGEIF(Daily!$A$2:$A${daily_last},"<>",Daily!$L$2:$L${daily_last}),0)', "₹#,##0.00")
    kpi_row(19, "EPB (avg of daily)", f'=IFERROR(AVERAGEIF(Daily!$A$2:$A${daily_last},"<>",Daily!$M$2:$M${daily_last}),0)', "₹#,##0")
    kpi_row(20, "Avg headway (min)", f'=IFERROR(AVERAGEIF(Daily!$A$2:$A${daily_last},"<>",Daily!$N$2:$N${daily_last}),0)', "0.0")

    ws["A22"] = "Gender (ticketed)"
    ws["A22"].font = SECTION_FONT
    kpi_row(
        23,
        "Male ridership",
        '=SUMIFS(Tickets[TP],Tickets[Gender],"M",Tickets[Date],">="&StartDate,Tickets[Date],"<="&EndDate)',
        "#,##0",
    )
    kpi_row(
        24,
        "Female ridership",
        '=SUMIFS(Tickets[TP],Tickets[Gender],"F",Tickets[Date],">="&StartDate,Tickets[Date],"<="&EndDate)',
        "#,##0",
    )

    ws["A26"] = "Sanity checks"
    ws["A26"].font = SECTION_FONT
    kpi_row(27, "Tickets in period", '=COUNTIFS(Tickets[Date],">="&StartDate,Tickets[Date],"<="&EndDate)', "#,##0")
    kpi_row(28, "Trips with capacity=0", f'=COUNTIFS(TripSummary!$J$2:$J${TRIP_ROWS+1},0,TripSummary!$A$2:$A${TRIP_ROWS+1},"<>")', "#,##0")
    kpi_row(29, "Trips with route len=0", f'=COUNTIFS(TripSummary!$I$2:$I${TRIP_ROWS+1},0,TripSummary!$A$2:$A${TRIP_ROWS+1},"<>")', "#,##0")

    ws["D4"] = "Matches Overview StatCards + Ops snapshot"
    ws["D4"].font = HINT_FONT
    ws["D5"] = "LF uses summed pax-km / capacity-km"
    ws["D5"].font = HINT_FONT
    ws["D14"] = "Denominator = sum of daily buses (bus-days)"
    ws["D14"].font = HINT_FONT
    ws["D18"] = "Dashboard uses mean of daily EPKM/EPB from kpi_daily"
    ws["D18"].font = HINT_FONT
    ws["D28"] = "If >0, fix Veh / Routes lookups"
    ws["D28"].fill = WARN_FILL

    autosize(ws, {1: 28, 2: 16, 3: 4, 4: 55})


def main() -> None:
    wb = Workbook()
    build_readme(wb)
    build_controls(wb)
    build_veh(wb)
    build_routes(wb)
    build_tickets(wb)
    build_trip_summary(wb)
    build_daily(wb)
    build_route_day(wb)
    build_kpi(wb)

    # Sensible tab order already created; freeze headers
    for name in ("Tickets", "TripSummary", "Daily", "RouteDay", "Veh", "Routes"):
        wb[name].freeze_panes = "A2"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
